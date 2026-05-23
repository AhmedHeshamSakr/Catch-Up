from openpyxl import load_workbook

from app.core.domain import (
    Category, DigestRun, Entity, Importance, NewsItem, RawItem, Sentiment, SourceType,
)
from app.services.render import excel


def _item(title, cat, *, summary="A summary.", imp=Importance.HIGH):
    raw = RawItem(source_id="s", source_type=SourceType.RSS, source_name="Src",
                  url="https://a.com/x", title=title, category_hint=cat)
    it = NewsItem.from_raw(raw, run_id="r1")
    it.category = cat
    it.summary_en = summary
    it.summary_ar = "ملخص"
    it.importance = imp
    it.entities = [Entity(name="OpenAI", type="org")]
    it.sentiment = Sentiment.NEUTRAL
    return it


def test_write_excel_has_master_and_category_sheets(tmp_path):
    run = DigestRun(run_id="rX")
    items = [_item("AI thing", Category.AI_TECH), _item("Gulf thing", Category.GULF_MENA)]
    path = excel.write_excel(run, items, str(tmp_path))
    assert path.endswith("digest-rX.xlsx")

    wb = load_workbook(path)
    assert "All News" in wb.sheetnames
    assert "AI & Technology" in wb.sheetnames
    assert "Gulf & MENA" in wb.sheetnames

    master = wb["All News"]
    assert [c.value for c in master[1]] == excel.HEADERS
    titles = [master.cell(row=r, column=2).value for r in range(2, master.max_row + 1)]
    assert "AI thing" in titles and "Gulf thing" in titles

    ai_sheet = wb["AI & Technology"]
    ai_titles = [ai_sheet.cell(row=r, column=2).value for r in range(2, ai_sheet.max_row + 1)]
    assert ai_titles == ["AI thing"]  # only AI items on that sheet


def test_excel_row_maps_all_fields(tmp_path):
    run = DigestRun(run_id="rY")
    path = excel.write_excel(run, [_item("Headline", Category.AI_TECH)], str(tmp_path))
    row = list(load_workbook(path)["All News"][2])
    values = [c.value for c in row]
    assert "Headline" in values
    assert "A summary." in values
    assert "HIGH" in values
    assert "OpenAI" in values
