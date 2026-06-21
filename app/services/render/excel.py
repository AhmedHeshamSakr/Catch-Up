from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.domain import DigestRun, NewsItem
from app.services.render.markdown import CATEGORY_TITLES

HEADERS = [
    "Date", "Title", "Summary (EN)", "Summary (AR)", "Category",
    "Source", "URL", "Importance", "Entities", "Sentiment",
]
_WIDTHS = [18, 50, 60, 60, 20, 18, 40, 12, 30, 12]
_HEADER_FILL = PatternFill("solid", fgColor="0F172A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Excel/Sheets interpret a cell starting with any of these as a formula → CSV/
# formula injection. We prefix a leading apostrophe to force plain text.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value: str) -> str:
    if value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _row(item: NewsItem) -> list[str]:
    when = item.published_at or item.collected_at
    category = CATEGORY_TITLES.get(item.category, "Uncategorized") if item.category else "Uncategorized"
    return [
        when.strftime("%Y-%m-%d %H:%M") if when else "",
        _safe_cell(item.title),
        _safe_cell(item.summary_en or ""),
        _safe_cell(item.summary_ar or ""),
        category,
        _safe_cell(item.source_name),
        _safe_cell(item.url),
        item.importance.value.upper() if item.importance else "",
        _safe_cell(", ".join(e.name for e in item.entities)),
        item.sentiment.value if item.sentiment else "",
    ]


def _write_sheet(ws: Worksheet, items: list[NewsItem]) -> None:
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for item in items:
        ws.append(_row(item))
    ws.freeze_panes = "A2"
    for i, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def render_workbook(run: DigestRun, items: list[NewsItem]) -> Workbook:
    wb = Workbook()
    master = wb.active
    master.title = "All News"
    _write_sheet(master, items)
    for category, title in CATEGORY_TITLES.items():
        group = [i for i in items if i.category == category]
        if group:
            _write_sheet(wb.create_sheet(title=title), group)
    return wb


def write_excel(run: DigestRun, items: list[NewsItem], output_dir: str) -> str:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"digest-{run.run_id}.xlsx"
    render_workbook(run, items).save(str(path))
    return str(path)
